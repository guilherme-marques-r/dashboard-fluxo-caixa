"""Acesso aos dados do fluxo de caixa (Google Sheets principal, xlsx local como fallback).

A fonte da verdade é o Google Sheets (aba Lançamentos), sempre atual pois o bot
grava os lançamentos lá. O xlsx local (cópia limpa) é usado como referência no
diagnóstico e como fallback quando não há acesso ao Google (modo offline).

Estrutura da aba Lançamentos (dados a partir da linha 3):
    A = Data (serial ou ISO)
    B = Tipo (Entrada/Saida)
    C = Valor
    D = Categoria nível 1
    E = Categoria nível 2
    F = Observação
    G = Subtotal (saldo corrente)
    E1 = saldo inicial, G1 = saldo atual (caixa)
"""
from datetime import datetime, timedelta

import gspread
import pandas as pd

from config import carregar_config

# Dia zero do Excel/Sheets (serial = dias desde 30/12/1899)
BASE_SERIAL = datetime(1899, 12, 30)

CONFIG = carregar_config()

COLUNAS = ["data", "tipo", "valor", "categoria_n1", "categoria_n2", "observacao", "subtotal"]


def obter_cliente():
    """Retorna cliente gspread autenticado com a service account.

    Prioriza a credencial fornecida via secrets do Streamlit Cloud (dict) e usa
    o JSON da service account do bot como alternativa local.
    """
    cred_dict = CONFIG.get("credencial_dict")
    if cred_dict:
        gc = gspread.service_account_from_dict(cred_dict)
        return gc
    if not CONFIG["credencial"]:
        raise FileNotFoundError("Credencial da service account não encontrada.")
    gc = gspread.service_account(filename=str(CONFIG["credencial"]))
    return gc


def listar_abas_sheets():
    """Lista os nomes das abas disponíveis no Google Sheets."""
    gc = obter_cliente()
    planilha = gc.open_by_key(CONFIG["planilha_id"])
    return [aba.title for aba in planilha.worksheets()]


def _normalizar_valor(valor):
    """Converte valor para float seguro.

    O Google Sheets retorna alguns campos já formatados como texto de moeda
    brasileira (ex.: ' R$ 130,00 ', ' R$ (122.466,98)'). Este parser lida com:
      - números puros (int/float)
      - texto com 'R$', espaços e separador de milhar
      - vírgula decimal (formato BR) ou ponto decimal (formato US)
      - negativo entre parênteses -> negativo
    """
    if valor is None:
        return 0.0
    if isinstance(valor, bool):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        return _parse_moeda(valor)
    try:
        return float(valor)
    except (TypeError, ValueError):
        return 0.0


def _parse_moeda(texto: str) -> float:
    """Interpreta uma string de moeda (pt-BR ou US) como float."""
    t = texto.strip().upper()
    if not t:
        return 0.0
    # Negativo entre parênteses: (1.234,56) -> -1234.56
    negativo = "(" in t and ")" in t
    t = t.replace("(", "").replace(")", "")
    # Remove rótulo de moeda, espaços e letras avulsas
    t = t.replace("R$", "").replace("$", "").replace("US", "").strip()
    t = t.replace(" ", "").replace("\u00a0", "")
    if not t:
        return 0.0
    # Se tiver vírgula, é o separador decimal pt-BR; pontos são milhar.
    if "," in t and t.count(",") == 1:
        t = t.replace(".", "").replace(",", ".")
    elif "." in t:
        # Sem vírgula: se houver um único ponto e após ele houver dígitos,
        # tratamos heuristicamente — exceto quando há múltiplos pontos (milhar).
        partes = t.split(".")
        if len(partes) == 2 and len(partes[1]) in (1, 2):
            t = t  # ponto decimal (ex.: '130.5' ou '130.50')
        else:
            t = t.replace(".", "")  # pontos como milhar
    try:
        valor = float(t)
    except ValueError:
        return 0.0
    return -valor if negativo else valor


def _serial_para_data(valor):
    """Converte serial do Excel/Sheets (dias desde 30/12/1899) para datetime."""
    if valor is None:
        return None
    try:
        numero = float(valor)
    except (TypeError, ValueError):
        return None
    if not (1000 <= numero <= 80000):  # faixa plausível de serial
        return None
    return BASE_SERIAL + timedelta(days=numero)


def _converter_data(valor):
    """Normaliza data vinda do Sheets: aceita serial, ISO, timestamp ou string."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        return _serial_para_data(valor)
    if isinstance(valor, datetime):
        return valor
    # String: tenta formatos comuns (ISO, dd/mm/aaaa)
    texto = str(valor).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d/%m/%Y %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(texto, fmt)
        except ValueError:
            continue
    try:
        return pd.to_datetime(texto, errors="raise")
    except Exception:
        return None


def ler_lancamentos_sheets():
    """Lê a aba Lançamentos do Google Sheets e devolve um DataFrame."""
    gc = obter_cliente()
    ws = gc.open_by_key(CONFIG["planilha_id"]).worksheet(CONFIG["aba_lancamentos"])
    # Linha 1 = controle, linha 2 = cabeçalho, dados a partir da linha 3.
    valores = ws.get("A3:G", value_render_option="UNFORMATTED_VALUE")
    linhas = []
    for linha in valores:
        if len(linha) < 7:
            linha = linha + [None] * (7 - len(linha))
        data = _converter_data(linha[0])
        if data is None:
            continue
        linhas.append({
            "data": data,
            "tipo": str(linha[1] or "").strip(),
            "valor": _normalizar_valor(linha[2]),
            "categoria_n1": str(linha[3] or "").strip(),
            "categoria_n2": str(linha[4] or "").strip(),
            "observacao": str(linha[5] or "").strip(),
            "subtotal": _normalizar_valor(linha[6]),
        })
    df = pd.DataFrame(linhas, columns=COLUNAS)
    df["data"] = pd.to_datetime(df["data"])
    return df


def ler_saldos_sheets():
    """Lê saldo inicial (E1) e saldo atual (G1) da aba Lançamentos."""
    gc = obter_cliente()
    ws = gc.open_by_key(CONFIG["planilha_id"]).worksheet(CONFIG["aba_lancamentos"])
    linha1 = ws.get("A1:G1")[0]
    saldo_inicial = _normalizar_valor(linha1[4]) if len(linha1) > 4 else None
    saldo_atual = _normalizar_valor(linha1[6]) if len(linha1) > 6 else None
    return {"saldo_inicial": saldo_inicial, "saldo_atual": saldo_atual}


def ler_lancamentos_xlsx():
    """Lê a aba Lançamentos do xlsx local (referência/fallback)."""
    from openpyxl import load_workbook

    caminho = CONFIG["xlsx_referencia"]
    wb = load_workbook(str(caminho), data_only=True, read_only=True)
    ws = wb[CONFIG["aba_lancamentos"]]
    linhas = []
    for row in ws.iter_rows(min_row=3, max_col=7, values_only=True):
        if row[0] is None:
            continue
        linhas.append({
            "data": row[0],
            "tipo": str(row[1] or "").strip(),
            "valor": _normalizar_valor(row[2]),
            "categoria_n1": str(row[3] or "").strip(),
            "categoria_n2": str(row[4] or "").strip(),
            "observacao": str(row[5] or "").strip(),
            "subtotal": _normalizar_valor(row[6]),
        })
    wb.close()
    df = pd.DataFrame(linhas, columns=COLUNAS)
    df["data"] = pd.to_datetime(df["data"], errors="coerce")
    df = df.dropna(subset=["data"])
    return df


def ler_lancamentos(fonte="sheets"):
    """Lê lançamentos. fonte = 'sheets' (padrão) ou 'xlsx' (offline)."""
    if fonte == "sheets":
        try:
            return ler_lancamentos_sheets()
        except Exception:
            if CONFIG["xlsx_referencia"].exists():
                return ler_lancamentos_xlsx()
            raise
    return ler_lancamentos_xlsx()


def ler_saldos(fonte="sheets"):
    """Lê saldos. fonte = 'sheets' (padrão) ou 'xlsx'."""
    if fonte == "xlsx":
        df = ler_lancamentos_xlsx()
        # Em datas reais primeiro/último subtotal não é confiável; usamos
        # o saldo do próprio arquivo quando disponível.
        return {"saldo_inicial": None, "saldo_atual": None}
    return ler_saldos_sheets()


def saldo_disponivel(df, saldo_inicial=None):
    """Calcula o saldo/caixa a partir dos lançamentos.

    Se saldo_inicial não for informado, usa o primeiro subtotal da coluna G.
    """
    if saldo_inicial is None or saldo_inicial == 0:
        primeira = df[df["subtotal"].notna()]["subtotal"]
        saldo_inicial = float(primeira.iloc[0]) if not primeira.empty else 0.0
    fluxo = df.apply(lambda r: r["valor"] if r["tipo"].lower().startswith("entrada") else -r["valor"], axis=1).sum()
    return saldo_inicial + fluxo


def _periodo_para_mes(valor):
    """Converte um período (string com mês/ano, data ou serial) em 'Aaaa-mm'."""
    if valor is None:
        return None
    import re

    texto = str(valor).strip()
    m = re.match(r"^(\d{1,2})[\/\-](\d{4})$", texto)  # 10/2026
    if m:
        return f"{m.group(2)}-{int(m.group(1)):02d}"
    m = re.match(r"^(\d{4})-(\d{2})", texto)  # 2026-10
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    d = _serial_para_data(valor)
    if d is not None:
        return d.strftime("%Y-%m")
    return None


def ler_producao_mensal():
    """Lê a produção mensal (kg) combinando:
    - aba 'Análises' (linha 'TOTAL DESPESCADO (kg)') — meses 2025+;
    - aba 'Relatório 1' ('Produção mensal (kg)') — 2023.
    Tenta o Google Sheets e cai para o xlsx local se vazio/ausente.
    Devolve um dict { 'Aaaa-mm': kg }.
    """
    try:
        gc = obter_cliente()
        producao = {}
        for nome_aba, rotulo in (("Análises", "despescado"), ("Relatório 1", "produc")):
            try:
                ws = gc.open_by_key(CONFIG["planilha_id"]).worksheet(nome_aba)
            except Exception:
                continue
            valores = ws.get("A1:AZ", value_render_option="UNFORMATTED_VALUE")
            producao.update(_parse_matriz_producao(valores, rotulo))
        if producao:
            return producao
        return _ler_producao_xlsx()
    except Exception:
        return _ler_producao_xlsx()


def _parse_matriz_producao(valores, rotulo="produc"):
    """Extrai a produção mensal (kg) da matriz bruta de uma aba.

    Procura a linha cujo rótulo contém `rotulo` (sem acentos) e devolve
    { 'Aaaa-mm': kg }, ignorando valores placeholder (0.001).
    """
    import unicodedata

    def sem_acento(texto):
        return unicodedata.normalize("NFD", str(texto)).encode("ascii", "ignore").decode("ascii").lower()

    if not valores:
        return {}
    cabecalho = list(valores[0]) + [None] * 60
    for linha in valores[1:]:
        if not linha or rotulo not in sem_acento(linha[0]):
            continue
        producao = {}
        for i in range(1, len(linha)):
            mes = _periodo_para_mes(cabecalho[i])
            kg = _normalizar_valor(linha[i])
            # 0.001 é placeholder usado na planilha para mês sem lançamento.
            if mes and kg > 1:
                producao[mes] = kg
        return producao
    return {}


def _ler_producao_xlsx():
    """Lê a produção mensal do xlsx local (Relatório 1 + aba Análises)."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(CONFIG["xlsx_referencia"]), data_only=True, read_only=True)
        resultado = {}
        for nome_aba, rotulo in (("Relatório 1", "produc"), ("Análises", "despescado")):
            if nome_aba not in wb.sheetnames:
                continue
            ws = wb[nome_aba]
            matriz = list(ws.iter_rows(max_row=10, values_only=True))
            resultado.update(_parse_matriz_producao(matriz, rotulo))
        wb.close()
        return resultado
    except Exception:
        return {}